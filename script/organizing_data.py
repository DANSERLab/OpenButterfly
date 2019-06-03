# Script to search differnt data files, grab relevant time frames, then consolidate them into one excel work book where each sheet is a differnt trial.
# requires this file to be at the top of the directory with all the other data folders
# the directory path should be of "DATA/Session#'s/Userfiles"



import os
import os.path
import copy
import csv
#import xlsxwriter module --> may need to download pip then run the command "pip install xlsxwriter"
import xlsxwriter
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import pandas as pd 
import re
import fnmatch

#============================================================================================================================
#					Variable Block
#============================================================================================================================


subject="Paniz"						# Subject's name as formated in File name
session="Session2"					# Session number to check
clock_start_times = []				# Array of global clock times that were recorded
hr_clock_start_time = []			# Time the HR monitor began recording
hr_offset_min = []
hr_offset_sec = []
hr_offset_start = []
hr_offset_end = []
hr_start_times = []					# Array of all the beginning times to clip
hr_end_times = []					# Array of all the end times to clip
gsr_start_times = []				# Array of all the beginning times to clip'
gsr_end_times_corrected = ["'0:1:0"]	# Initialize with baseline values since it't just the first 60 sec
gsr_start_times_corrected = ["'-0:0:0"]
date = []							# Date of session
# number_of_recordings = []			# The number of Muse/Neulog recordings
recording_transitions = ['1']			# Array for each "part".  Shows transitions from recording 1 to recording 2 to recording 3, etc
recording_transitions_muse = []
muse_start_time = []				# Array of all the begining times to clip
muse_start_times = []
muse_end_times = []
array_game = []
file_array_muse = []
file_array_gsr = []
muse_min_array = []
muse_hour_array = []
muse_sec_array = []



ex_list_session = []
ex_list_session1to6 = ['FAR', 'SAR', 'SR','FAR', 'SAR', 'SR','FAR', 'SAR', 'SR']
ex_list_session7to12 = ['ExR', 'AbR', 'MxdPr', 'MxdCr', 'ExR', 'AbR', 'MxdPr', 'MxdCr', 'FAR', 'SAR', 'SR','FAR', 'SAR', 'SR']

trial_list_session = []
trial_list_session1to6 = ['Trial1', 'Trial1', 'Trial1', 'Trial2', 'Trial2', 'Trial2', 'Trial3', 'Trial3', 'Trial3']
tiral_list_session7to12 = ['Trial1','Trial1','Trial1','Trial1','Trial2', 'Trial2', 'Trial2', 'Trial2','Trial1','Trial1','Trial1', 'Trial2', 'Trial2', 'Trial2']
############################################################################################################################	
#============================================================================================================================
#					Function Block
#============================================================================================================================
#############################################################################################################################	
#============================================================================================================================

# FUNCTION:	Make excel workbook with correct name and sheet names
#			This will over write any workbook with the same name.

def make_excel_book(name,session):
	global clip_file_name

	if session in ["Session1", "Session2", "Session3", "Session4", "Session5", "Session6"]:
		# Workbook() takes one, non-optional, argument  
		# which is the filename that we want to create. 
		workbook = xlsxwriter.Workbook('clipped_' + name + '_' + session + '.xlsx') 
		  
		# By default worksheet names in the spreadsheet will be  
		# Sheet1, Sheet2 etc., but we can also specify a name. 
		worksheet = workbook.add_worksheet("baseline") 
		worksheet = workbook.add_worksheet("FAR1") 
		worksheet = workbook.add_worksheet("SAR1") 
		worksheet = workbook.add_worksheet("SR1") 
		worksheet = workbook.add_worksheet("FAR2") 
		worksheet = workbook.add_worksheet("SAR2") 
		worksheet = workbook.add_worksheet("SR2") 
		worksheet = workbook.add_worksheet("FAR3") 
		worksheet = workbook.add_worksheet("SAR3") 
		worksheet = workbook.add_worksheet("SR3") 
		

		clip_file_name = 'clipped_' + name + '_' + session + '.xlsx'
		#sheet_names = ["baseline","FAR1","SAR1", "SR1", "FAR2", "SAR2", "SR2", "FAR3", "SAR3", "SR3"]
		# Finally, close the Excel file 
		# via the close() method. 
		workbook.close() 

	elif session in ["Session7","Session8","Session9","Session10","Session11","Session12"]:
		# Workbook() takes one, non-optional, argument  
		# which is the filename that we want to create. 
		workbook = xlsxwriter.Workbook('clipped_' + name + '_' + session + '.xlsx') 
		  
		# By default worksheet names in the spreadsheet will be  
		# Sheet1, Sheet2 etc., but we can also specify a name. 
		worksheet = workbook.add_worksheet("baseline") 
		worksheet = workbook.add_worksheet("ExR1") 
		worksheet = workbook.add_worksheet("AbR1") 
		worksheet = workbook.add_worksheet("MxdPr1") 
		worksheet = workbook.add_worksheet("MxdCr1") 
		worksheet = workbook.add_worksheet("ExR2") 
		worksheet = workbook.add_worksheet("AbR2") 
		worksheet = workbook.add_worksheet("MxdPr2") 
		worksheet = workbook.add_worksheet("MxdCr2")
		worksheet = workbook.add_worksheet("FAR1") 
		worksheet = workbook.add_worksheet("SAR1") 
		worksheet = workbook.add_worksheet("SR1") 
		worksheet = workbook.add_worksheet("FAR2") 
		worksheet = workbook.add_worksheet("SAR2") 
		worksheet = workbook.add_worksheet("SR2") 
		
		clip_file_name = 'clipped_' + name + '_' + session + '.xlsx'  
		#sheet_names = ["baseline", "ExR1", "AbR1", "MxdPr1", "MxdCr1", "ExR2", "AbR2", "MxdPr2", "MxdCr2", "FAR1","SAR1", "SR1", "FAR2", "SAR2", "SR2"]
		# Finally, close the Excel file 
		# via the close() method. 
		workbook.close() 

	else:
			print("This is not a valid session name. Should be Session1, Session2,...,Session12.  No Session0 or Session1")


#############################################################################################################################	
#============================================================================================================================

# FUNCTION:  Start Times ./start_time/Session1/time_Mike.csv
	#	make array of start times heart rate, gsr, and muse
	# csv file with column 0 of HR,Muse/GSR,Baseline,Trial1, Trial 2... 9, Survey, ROM
	#				column 1 of Clock time of the form HOUR:MIN
	#				column 2 of GSR time
def make_start_time_array(name,session):
	global trial_list_session
	global ex_list_session

	with open(file_time, 'rb') as csvFile:
    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
    		originalFile_time = list(reader)						# make a list of spread sheet, need a list to index to a specfic cell and overwrite old spread sheet
	
		# get number of recordings
		global number_of_recordings
		date.append(originalFile_time[2][1])
		number_of_recordings = int(originalFile_time[3][1])

		# this section builds the gsr start times, rec transitions and exercise/trial list based on the session that is being examined
		if session in ["Session1","Session2","Session3","Session4","Session5","Session6"]:
			i = 8
			while i <= 17:
				gsr_start_times.append(originalFile_time[i][2])
				recording_transitions.append(originalFile_time[i][3])
				recording_transitions_muse.append(originalFile_time[i][3])  
				i += 1

			j = 5
			while j <= 18:
				clock_start_times.append(originalFile_time[j][1]) 
				j += 1

			ex_list_session = ['FAR', 'SAR', 'SR','FAR', 'SAR', 'SR','FAR', 'SAR', 'SR']
			trial_list_session = ['Trial1', 'Trial1', 'Trial1', 'Trial2', 'Trial2', 'Trial2', 'Trial3', 'Trial3', 'Trial3']


 		elif session in ["Session7","Session8","Session9","Session10","Session11","Session12"]:
			# make speciic cells of a column into an array
			i = 8
			while i <= 22:
				gsr_start_times.append(originalFile_time[i][2])
				recording_transitions.append(originalFile_time[i][3]) 
				recording_transitions_muse.append(originalFile_time[i][3])
				i += 1

			j = 5
			while j <= 23:
				clock_start_times.append(originalFile_time[j][1]) 
				j += 1

			ex_list_session = ['ExR', 'AbR', 'MxdPr', 'MxdCr', 'ExR', 'AbR', 'MxdPr', 'MxdCr', 'FAR', 'SAR', 'SR','FAR', 'SAR', 'SR']
			trial_list_session = ['Trial1','Trial1','Trial1','Trial1','Trial2', 'Trial2', 'Trial2', 'Trial2','Trial1','Trial1','Trial1', 'Trial2', 'Trial2', 'Trial2']
			
	 	csvFile.close()


#############################################################################################################################			
#============================================================================================================================

# FUNCTION: Heart Rate ./heart_rate/Session1/HR_Mike.csv
def heart_rate_clock_start_time(name,session):

	# First get the global start time for the Heart Rate Recording and break into hours,min, and sec
	global originalFile_HR
	with open(file_HR, 'rb') as csvFile:
    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
    		originalFile_HR = list(reader)							# make a list of spread sheet, need a list to index to a specfic cell and overwrite old spread sheet
	
	hr_clock_start_time.append(originalFile_HR[1][3])

	global hr_hour, hr_min, hr_sec
	hr_hour, hr_min, hr_sec = hr_clock_start_time[0].split(":")
	hr_hour = int(hr_hour)
	hr_min = int(hr_min)
	hr_sec = int(hr_sec)
	
	csvFile.close()

############################################################################################################################		
#============================================================================================================================

# FUNCTION: find the offset from beginning of HR file to beginning of particular Muse file
def hr_find_offsets(input_file):
	# Next get start minute of Muse which should be later than the Heart rate.  
	# Need to take care of cases where there are multiple muse recordings
	global originalFile_muse
	with open(input_file, 'rb') as csvFile:
    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
    		originalFile_muse = list(reader)	

    	muse_start_time.append(originalFile_muse[1][0])

    	# break the muse start time into hour, min and seconds
    	global muse_date, muse_hour, muse_min, muse_sec, muse_decisec

    	muse_hour, muse_min, muse_sec = muse_start_time[-1].split(":")
    	muse_date, muse_hour = muse_hour.split(" ")
    	muse_sec, muse_decisec = muse_sec.split(".")
    	muse_hour = int(muse_hour)
    	muse_min = int(muse_min)
    	muse_sec = int(muse_sec)
    	muse_decisec = int(muse_decisec)

    	muse_hour_array.append(muse_hour)
    	muse_min_array.append(muse_min)
    	muse_sec_array.append(muse_sec)

    	csvFile.close()

    	temp_hour = muse_hour
    	temp_min = muse_min+1
    	if temp_min>59:
    		temp_min = temp_min - 60
    		temp_hour = temp_hour + 1

    	# get the initial start time for the baseline
    	if not muse_start_times:
    			if muse_sec > 9: t_muse_sec = str(muse_sec)
    			if muse_sec < 10: t_muse_sec = "0"+str(muse_sec)
			if muse_min > 9: t_muse_min = str(muse_min)
			if muse_min < 10: t_muse_min = '0' + str(muse_min)
			if temp_min > 9: t_temp_min = str(temp_min)
			if temp_min < 10: t_temp_min = '0' + str(temp_min)
			if muse_hour > 9: t_muse_hour = str(muse_hour)
			if muse_hour < 10: t_muse_hour = '0' + str(muse_hour)
			if temp_hour > 9: t_temp_hour = str(temp_hour)
			if temp_hour < 10: t_temp_hour = '0' + str(temp_hour)

	    		muse_start_times.append(t_muse_hour + ':' + t_muse_min + ':' + t_muse_sec)
	    		muse_end_times.append(t_temp_hour + ':' + t_temp_min + ':' + t_muse_sec)

	    	# muse_start_times.append(str(muse_hour) + ':' + str(muse_min) + ':' + str(muse_sec))
	    	# muse_end_times.append(str(temp_hour) + ':' + str(temp_min) + ':' + str(muse_sec))


    	# If both muse and HR hours match then just subtract HR min and sec from Muse min and sec to get offse
    	# If muse and HR hours differ subtract HR min and sec from 60 then add to Muse min and sec

    	if hr_hour == muse_hour:
    		diff_min = muse_min - (hr_min + 1)
    		diff_sec = muse_sec + (60-hr_sec)
    		if diff_sec > 59:
    			diff_min = diff_min + 1
    			diff_sec = diff_sec - 60

    		hr_offset_min.append(diff_min)
    		hr_offset_sec.append(diff_sec)


    	if hr_hour!=muse_hour:
    		diff_min = (60 - (hr_min + 1)) + muse_min
    		diff_sec = (60 - hr_sec) + muse_sec
    		if diff_sec > 59:
    			diff_min = diff_min + 1
    			diff_sec = diff_sec - 60

    		hr_offset_min.append(diff_min)
    		hr_offset_sec.append(diff_sec)

    	
############################################################################################################################	
#============================================================================================================================

# FUNCTION:  Need to add the offsets to each of the times from gsr_start_times to get hr_end_time
def start_end_times():
	
	#local arrays to put in "clock" format
	time_temp_hr = []		
	time_temp_gsr = []

	for i in range(len(hr_offset_start)):
		time_temp_hr.append(datetime.strptime(hr_offset_start[i], '%H:%M:%S').time())

	for j in range(len(gsr_start_times)):							# sometimes data is missing, we fill with an "X" to account for this
		if gsr_start_times[j] == 'X': time_temp_gsr.append('X')
		else:
			time_temp_gsr.append(datetime.strptime(gsr_start_times[j], '%H:%M:%S').time())
	# print(gsr_start_times)


	# This large for loop will create with start and stop arrays in the correct format, accounts for sec sum>59, or if time is missing indicated by 'X'
	for k in range(len(gsr_start_times)):
		offset_index = int(recording_transitions[k]) - 1
		rec = int(recording_transitions_muse[k])-1
		if gsr_start_times[k] == 'X': 
			hr_start_times.append('X')
			hr_end_times.append('X')
			gsr_start_times_corrected.append('X')
			gsr_end_times_corrected.append('X')
			muse_start_times.append('X')
			muse_end_times.append('X')
		else:
			sec_sum = int(time_temp_hr[offset_index].second)+int(time_temp_gsr[k].second)
			min_sum = int(time_temp_hr[offset_index].minute)+int(time_temp_gsr[k].minute)
			min_sum_end = min_sum + 1
			sec_sum_gsr = int(time_temp_gsr[k].second)
			min_sum_gsr = int(time_temp_gsr[k].minute)
			min_sum_end_gsr = min_sum_gsr + 1
			sec_sum_muse = int(time_temp_gsr[k].second) + muse_sec_array[rec]
			min_sum_muse = int(time_temp_gsr[k].minute) + muse_min_array[rec]
			min_sum_end_muse = min_sum_muse + 1
			hour_sum_muse = muse_hour_array[rec]
			hour_sum_end_muse = muse_hour_array[rec]
			print(time_temp_gsr[k].minute, muse_min_array[rec], min_sum_muse,recording_transitions_muse[k])
			if sec_sum > 59: 
				sec_sum = sec_sum - 60
				min_sum = min_sum + 1
				min_sum_end = min_sum_end + 1
			if sec_sum_gsr > 59:
				sec_sum_gsr = sec_sum_gsr - 60
				min_sum_gsr = min_sum_gsr + 1
				min_sum_end_gsr = min_sum_end_gsr + 1
			if sec_sum_muse > 59: 
				sec_sum_muse = sec_sum_muse - 60
				min_sum_muse = min_sum_muse + 1
				min_sum_end_muse = min_sum_end_muse + 1
			if min_sum_muse > 59: 
				min_sum_muse = min_sum_muse - 60
				hour_sum_muse = hour_sum_muse + 1
			if min_sum_end_muse >59:
				min_sum_end_muse = min_sum_end_muse - 60
				hour_sum_end_muse = hour_sum_end_muse + 1
			sec_sum_gsr = str(sec_sum_gsr)
			min_sum_gsr = str(min_sum_gsr)
			min_sum_end_gsr = str(min_sum_end_gsr)
			hour_sum_muse = str(hour_sum_muse)
			hour_sum_end_muse = str(hour_sum_end_muse)
			if sec_sum > 9: sec_sum = str(sec_sum)
			if sec_sum < 10: sec_sum = "0"+str(sec_sum)
			if min_sum > 9: min_sum = str(min_sum)
			if min_sum < 10: min_sum = '0' + str(min_sum)
			if min_sum_end > 9: min_sum_end = str(min_sum_end)
			if min_sum_end < 10: min_sum_end = '0' + str(min_sum_end)
			if sec_sum_muse > 9: sec_sum_muse = str(sec_sum_muse)
			if sec_sum_muse < 10: sec_sum_muse = "0"+str(sec_sum_muse)
			if min_sum_muse > 9: min_sum_muse = str(min_sum_muse)
			if min_sum_muse < 10: min_sum_muse = '0' + str(min_sum_muse)
			if min_sum_end_muse > 9: min_sum_end_muse = str(min_sum_end_muse)
			if min_sum_end_muse < 10: min_sum_end_muse = '0' + str(min_sum_end_muse)
			# print(min_sum_muse)

			#Creating each start and stop time arrays that are used to search for the correct rows to start and stop at
			hr_start_times.append('00:' + min_sum + ':' +  sec_sum)
			hr_end_times.append('00:' + min_sum_end + ':' + sec_sum)

			gsr_start_times_corrected.append("'0:" + min_sum_gsr + ':' + sec_sum_gsr)
			gsr_end_times_corrected.append("'0:" + min_sum_end_gsr + ':' + sec_sum_gsr)

			if k != 0:
				muse_start_times.append(hour_sum_muse + ':' + min_sum_muse + ':' + sec_sum_muse)
				muse_end_times.append(hour_sum_end_muse + ':' + min_sum_end_muse + ':' + sec_sum_muse)

	print(muse_start_times)
	print(len(muse_start_times),len(gsr_start_times))



############################################################################################################################	
#============================================================================================================================

# FUNCTION:  append the correct heart rate times to each sheet
def clip_hr():
	# flag is high after start time and low after end time
	clip_flag = 0

	wb = openpyxl.load_workbook(clip_file_name)

	for h in range(len(hr_start_times)-1):
		for i, row in enumerate(originalFile_HR):
			if i == 2: 	wb.worksheets[h].append(row)		
			for field in row:  
				if hr_start_times[h] in field: clip_flag = 1
				if hr_end_times[h] in field: clip_flag = 0
				if clip_flag == 1: 
					wb.worksheets[h].append(row)

	wb.save(clip_file_name)

############################################################################################################################	
#============================================================================================================================

# FUNCTION:	This function clips the appropriate rows of the gsr column to the correct sheet of the excel work book
def clip_gsr():
	global originalFile_gsr
	with open(file_gsr, 'rU') as csvFile:
    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
    		originalFile_gsr = list(reader)
    	csvFile.close()



	clip_flag = 0 
	wb = openpyxl.load_workbook(clip_file_name)

	
	row_count = 2

	for h in range(len(gsr_start_times_corrected)-1):
		if recording_transitions[h] == '1':
			wb.worksheets[h].cell(row=1, column=4).value =	"Time_gsr"
			wb.worksheets[h].cell(row=1, column=5).value =	"value_gsr"
			for i, row in enumerate(originalFile_gsr):
				for j, col in enumerate(row):
					# if i == 7:  wb.worksheets[0].cell(row=i, column=j+4).value = originalFile_gsr[i][j]
					for field in row:
						#if gsr_start_times_corrected[2] in field:  print(i)
						if gsr_start_times_corrected[h] in field: clip_flag = 1
						if gsr_end_times_corrected[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+4).value = originalFile_gsr[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

		if recording_transitions[h] != '1':
			updated_file_gsr = os.path.join(fileDir, 'gsr/', session, 'gsr_'+ subject + '_part' + recording_transitions[h] + '.csv')
			with open(updated_file_gsr, 'rU') as csvFile:
		    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
		    		originalFile_gsr = list(reader)
		    	csvFile.close()

			clip_flag = 0 
			row_count = 2

			wb.worksheets[h].cell(row=1, column=4).value =	"Time_gsr"
			wb.worksheets[h].cell(row=1, column=5).value =	"value_gsr"
			for i, row in enumerate(originalFile_gsr):
				for j, col in enumerate(row):
					# if i == 7:  wb.worksheets[0].cell(row=i, column=j+4).value = originalFile_gsr[i][j]
					for field in row:
						#if gsr_start_times_corrected[2] in field:  print(i)
						if gsr_start_times_corrected[h] in field: clip_flag = 1
						if gsr_end_times_corrected[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+4).value = originalFile_gsr[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

	wb.save(clip_file_name)

	
############################################################################################################################	
#============================================================================================================================

# FUNCTION:	This function clips the appropriate rows of the muse to the correct sheet of the excel work book
def clip_muse():
	clip_flag = 0 
	wb = openpyxl.load_workbook(clip_file_name)

	
	row_count = 2

	for h in range(len(muse_start_times)):
		if recording_transitions_muse[h] == '1':
			for i, row in enumerate(originalFile_muse_1):
				for j, col in enumerate(row):
					if i == 1:  
						wb.worksheets[h].cell(row=1, column=j+6).value = originalFile_muse_1[0][j]
					for field in row:
						if muse_start_times[h] in field: clip_flag = 1
						if muse_end_times[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+6).value = originalFile_muse_1[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

		if recording_transitions_muse[h] == '2':
			for i, row in enumerate(originalFile_muse_2):
				for j, col in enumerate(row):
					if i == 1:  
						wb.worksheets[h].cell(row=1, column=j+6).value = originalFile_muse_2[0][j]
					for field in row:
						if muse_start_times[h] in field: clip_flag = 1
						if muse_end_times[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+6).value = originalFile_muse_2[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

		if recording_transitions_muse[h] == '3':
			for i, row in enumerate(originalFile_muse_3):
				for j, col in enumerate(row):
					if i == 1:  
						wb.worksheets[h].cell(row=1, column=j+6).value = originalFile_muse_3[0][j]
					for field in row:
						if muse_start_times[h] in field: clip_flag = 1
						if muse_end_times[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+6).value = originalFile_muse_3[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

		if recording_transitions_muse[h] == '4':
			for i, row in enumerate(originalFile_muse_4):
				for j, col in enumerate(row):
					if i == 1:  
						wb.worksheets[h].cell(row=1, column=j+6).value = originalFile_muse_4[0][j]
					for field in row:
						if muse_start_times[h] in field: clip_flag = 1
						if muse_end_times[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+6).value = originalFile_muse_4[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

		if recording_transitions_muse[h] == '5':
			for i, row in enumerate(originalFile_muse_1):
				for j, col in enumerate(row):
					if i == 1:  
						wb.worksheets[h].cell(row=1, column=j+6).value = originalFile_muse_5[0][j]
					for field in row:
						if muse_start_times[h] in field: clip_flag = 1
						if muse_end_times[h] in field: clip_flag = 0
						if clip_flag == 1: 
							wb.worksheets[h].cell(row=row_count, column=j+6).value = originalFile_muse_5[i][j]
				if clip_flag == 1:  row_count += 1
			row_count = 2

	wb.save(clip_file_name)


############################################################################################################################	
#============================================================================================================================

# FUNCTION: get game play data file paths and store in array
def game_file_paths():
	global array_game
	for i in range(len(trial_list_session)): 
		temp_path = fnmatch.filter(os.listdir(fileDir+'/game/'+session), ex_list_session[i] +'*'+subject+'*'+trial_list_session[i]+'*'+'.csv')
		
		if not temp_path:
			temp_path = "X"

		array_game.append(temp_path)

############################################################################################################################	
#============================================================================================================================

# FUNCTION:	This function clips the appropriate rows of the game file to the correct sheet of the excel work book

def clip_game():
	clip_flag = 1 
	wb = openpyxl.load_workbook(clip_file_name)
	check = '60.0'

	for i in range(len(array_game)):
		if array_game[i][0] != 'X':
			file_game = os.path.join(fileDir, 'game/', session, array_game[i][0])
			with open(file_game, 'rb') as csvFile:
		    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
		    		originalFile_game = list(reader)
		    	csvFile.close()
		
		row_count = 1
		clip_flag = 1

		print(i)

		for k, row in enumerate(originalFile_game):
			for j, col in enumerate(row):
				for time in row:
					if clip_flag == 1 and j < 25 and k<6000: 
						wb.worksheets[i+1].cell(row=row_count, column=j+45).value = originalFile_game[k][j]
					if check in time and k>1500: 
						clip_flag = 0
						break #clip_flag = 0
					#if '0.0' in field: clip_flag = 1					
			if clip_flag == 1 and k>2:  
				row_count += 1

			if clip_flag == 0:
				break
		
		row_count = 1


	wb.save(clip_file_name)






##############################################################################################################################	
##############################################################################################################################	
	           		# MAIN CODE BLOCK #
##############################################################################################################################	
##############################################################################################################################	


# Current working directory
fileDir = os.path.dirname(os.path.realpath('__file__'))
#For accessing the file in a folder contained in the current folder
file_time = os.path.join(fileDir, 'start_time/', session, 'time_'+ subject +'.csv')
file_gsr = os.path.join(fileDir, 'gsr/', session, 'gsr_'+ subject +'.csv')
file_HR = os.path.join(fileDir, 'heart_rate/', session, 'HR_'+ subject +'.csv')
file_muse = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject +'.csv')


make_excel_book(subject,session)

# take in subject and session 
#subject = raw_input("Enter subject's name to consolidate: ")
#session = input("Enter subject's session to consolidate (ie Sesion2, Session5): ")
#print ("You entered " + subject + "and " + session) 

make_start_time_array(subject,session)
heart_rate_clock_start_time(subject,session)

# Realized for mulitple recordings there are multiple parts for muse and gsr. Rather than changing a bunch of code right now going to make arrays 
# for the file paths.


# This for loop finds Heart Rate start time offsets for any number of recordings.
# Think there is an error here --> should be if recording number != 1 then run this loop for each recording 
for i in range(len(recording_transitions)): # should be length ofrecording transitions not number
	if i == 0:
		hr_find_offsets(file_muse)
		#file_array_muse.append()
	if i > 0 and recording_transitions[i] > recording_transitions[i-1]:
		file_muse_updated = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject + '_part' + recording_transitions[i] + '.csv')
		hr_find_offsets(file_muse_updated) # Need to make sure there are matching parts for each recording

# Loop to concatenate min and sec (accounts for numbers less than 10, so has correct number of digits).  By the end of an array of hr start times and end times to search for
for i in range(len(hr_offset_min)):
	if hr_offset_min[i] < 10:
		temp_min = '0' + str(hr_offset_min[i])
	if hr_offset_min[i] > 9:
		temp_min = str(hr_offset_min[i])
	if hr_offset_sec[i] < 10:
		temp_sec = '0' + str(hr_offset_sec[i])
	if hr_offset_sec[i] > 9:
		temp_sec = str(hr_offset_sec[i])
	hr_offset_start.append('00:' + temp_min + ':' + temp_sec)
	if hr_offset_min[i]<9:
		temp_min_end = '0' + str(hr_offset_min[i]+1)
	if hr_offset_min[i]>8:
		temp_min_end = str(hr_offset_min[i]+1)
	hr_offset_end.append('00:' + temp_min_end + ':' + temp_sec)
	if i == 0:	#This if statement does the baseline.  Appends just the initial offset to the heart rate start and end times
		hr_start_times.append(str(hr_offset_start[0]))
		hr_end_times.append(str(hr_offset_end[0]))

start_end_times()
clip_hr()
clip_gsr()

#	Need a loop to make a list for each muse file, the call clip muse and have it get thr correct list
for m in range(1,number_of_recordings+1):
	if m ==1:
		file_muse_1 = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject +'.csv')
		global originalFile_muse_1
		with open(file_muse_1, 'rb') as csvFile:
	    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
	    		originalFile_muse_1 = list(reader)

	if m ==2:
		file_muse_2 = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject + '_part' + str(m) + '.csv')
		global originalFile_muse_2
		with open(file_muse_2, 'rb') as csvFile:
	    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
	    		originalFile_muse_2 = list(reader)

	if m ==3:
		file_muse_3 = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject + '_part' + str(m) + '.csv')
		global originalFile_muse_3
		with open(file_muse_3, 'rb') as csvFile:
	    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
	    		originalFile_muse_3 = list(reader)

	if m ==4:
		file_muse_4 = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject + '_part' + str(m) + '.csv')
		global originalFile_muse_4
		with open(file_muse_4, 'rb') as csvFile:
	    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
	    		originalFile_muse_4 = list(reader)

	if m ==5:
		file_muse_5 = os.path.join(fileDir, 'muse/', session, 'muse_'+ subject + '_part' + str(m) + '.csv')
		global originalFile_muse_5
		with open(file_muse_5, 'rb') as csvFile:
	    		reader = csv.reader(csvFile)							# create reader wrapped around an object.  These means use one time and done, so can't call twice.
	    		originalFile_muse_5 = list(reader)




clip_muse()
# game_file_paths()
# clip_game()
# print(hr_offset_start)
# print(hr_offset_end)
# print(hr_start_times)
# print(number_of_recordings)
#print(array_game[2df][0])
# print(hr_offset_start)
# print(hr_start_times)
# print(hr_offset_start[0])
# print(gsr_start_times_corrected)
# print(recording_transitions)

# print(muse_start_times)
# print(muse_min_array)
print('gsr start times')
print(gsr_start_times_corrected)
print('gsr end times')
print(gsr_end_times_corrected)

